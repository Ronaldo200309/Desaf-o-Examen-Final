"""
Desafío 2: Generación y modificación de un archivo Excel con openpyxl

Descripción:
Crear un script que cargue un archivo Excel existente, agregue algunos datos de forma programática,
y luego guarde el archivo modificado.
"""

import openpyxl
import os
from openpyxl.utils import get_column_letter

def update_excel(file_path):
    if not file_path.endswith('.xlsx'):
        print("Error: El archivo debe tener la extensión '.xlsx'.")
        return
    
    if not os.path.exists(file_path):
        print("El archivo no existe. Creando un nuevo archivo...")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Nombre", "Nota", "Asignatura"])  
        wb.save(file_path)
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    
    while True:
        nombre = input("Ingrese el nombre (o escriba 'salir' para terminar): ")
        if nombre.lower() == 'salir':
            break
        nota = input("Ingrese la nota: ")
        asignatura = input("Ingrese la asignatura: ")
        
        new_row = [nombre, nota, asignatura]
        sheet.append(new_row)

        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column].width = max_length + 2
    
    try:
        wb.save(file_path)
        print("Los datos fueron agregados correctamente al archivo Excel.")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
    
    wb.close()
    
file_path = "notas.xlsx"
update_excel(file_path)
